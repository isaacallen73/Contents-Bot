from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Column spec: (header, width, item_field_or_None, number_format_or_None)
COLUMNS = [
    ('Photo #',        22, 'primary_photo', None),
    ('Category',       18, 'category',      None),
    ('Manufacturer',   18, 'manufacturer',  None),
    ('Item',           38, 'item',          None),
    ('Model/Serial #', 20, 'model_serial',  None),
    ('Quanity',        10, 'quantity',      None),   # sic — matches client template
    ('Price',          12, 'price',         '"$"#,##0.00'),
    ('Total',          12, None,            '"$"#,##0.00'),  # formula
    ('Age',            10, 'age',           None),
]

_THIN = Side(style='thin')
_HEADER_FONT = Font(name='Calibri', bold=True, underline='single', size=11)
_DATA_FONT = Font(name='Calibri', size=11)
_LINK_FONT = Font(name='Calibri', size=11, color='0563C1', underline='single')
_TOTAL_FONT = Font(name='Calibri', bold=True, size=11)


def export_to_excel(items: list, output_path: str, photo_folder: str = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    # Header row
    for col_idx, (header, width, _, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 18

    # Data rows
    for row_idx, item in enumerate(items, 2):
        for col_idx, (_, _, field, num_fmt) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = _DATA_FONT
            cell.alignment = Alignment(vertical='center')

            if col_idx == 8:  # Total — formula
                if item.get('price') is not None:
                    cell.value = f'=G{row_idx}*F{row_idx}'
                    cell.number_format = '"$"#,##0.00'
            elif field:
                val = item.get(field)
                cell.value = val if val not in (None, '') else None
                if num_fmt and val is not None:
                    cell.number_format = num_fmt

            # Photo # as hyperlink
            if col_idx == 1 and photo_folder and item.get('primary_photo'):
                photo_path = str(Path(photo_folder) / item['primary_photo'])
                try:
                    cell.hyperlink = photo_path
                    cell.font = _LINK_FONT
                except Exception:
                    pass

            # Price as hyperlink to source if available
            if col_idx == 7 and item.get('price_source_url'):
                try:
                    cell.hyperlink = item['price_source_url']
                    cell.font = _LINK_FONT
                except Exception:
                    pass

    # Total row
    total_row = len(items) + 2
    label_cell = ws.cell(row=total_row, column=7, value='Total')
    label_cell.font = _TOTAL_FONT
    label_cell.alignment = Alignment(horizontal='right')

    total_cell = ws.cell(row=total_row, column=8, value=f'=SUM(H2:H{total_row - 1})')
    total_cell.font = _TOTAL_FONT
    total_cell.number_format = '"$"#,##0.00'

    for col_idx in (7, 8, 9):
        ws.cell(row=total_row, column=col_idx).border = Border(top=_THIN)

    wb.save(output_path)
    return output_path
